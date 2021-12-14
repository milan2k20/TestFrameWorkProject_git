import csv
import inspect
import logging
import softest
from openpyxl import load_workbook


class Utils(softest.TestCase):
    def assert_list_item_text(self, arr, value):
        for element in arr:
            print('The text is:' + element.text)
            self.soft_assert(self.assertEqual, element.text, value)
            if element.text == value:
                print('test passed')
            else:
                print('test failed')
        self.assert_all()


    def custom_logger(logLevel=logging.DEBUG):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        file_handler_obj = logging.FileHandler(
            "E:\Python program\Python Practice\Python Logging\Log File\demo_log_file1.txt")
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s',
                                      datefmt='%d/%m/%Y %I:%M:%S %p')
        file_handler_obj.setFormatter(formatter)
        logger.addHandler(file_handler_obj)
        return logger


    def read_data_from_excelfile(file_name, sheet):
        data_list = []
        load_workbiik_obj = load_workbook(filename=file_name)
        sh = load_workbiik_obj[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column
        for x in range(2, row_ct + 1):
            row = []
            for y in range(1, col_ct + 1):
                row.append(sh.cell(row=x, column=y).value)
            data_list.append(row)
        return data_list

    def read_data_from_csvfile(file_name):
        data_list = []
        # open csv file
        csv_data = open(file_name, 'r')
        # create csv reader
        reader = csv.reader(csv_data)
        # skip header
        next(reader)
        for rows in reader:
            data_list.append(rows)
        return data_list